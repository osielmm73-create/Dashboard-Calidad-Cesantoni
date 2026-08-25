
Buscar conversaciones
Imágenes
Biblioteca
Nuevo cuaderno
Diseño de Dashboard Operativo
Añadir Flejes Verdes Horizontales a Palet
Eliminar flejes verdes verticales centrales
Alineación de Cuerdas Verdes en Cajas
Cajas de Porcelánico Cesantoni en Tarima
Rama • Rama • Replicar Estilo De Imagen En Otra
Rama • Replicar Estilo De Imagen En Otra
Replicar Estilo De Imagen En Otra
Imagen Similar Estilo Vectorial
Conversación con Gemini
quiero hacer un dashboard con la información está en las tablas siguientes 



Primer tabla 

A:D

Columna A MES CALIDAD ANUAL

Columna B P1 ANUAL

Columna C P3 ANUAL

Columna D P1&P3 ANUAL 



SEGUNDA TABLA 

F:J

Columna F DÍA CALIDAD DIARIA 

Columna G P1 CALIDAD DIARIA 

Columna H P3 CALIDAD DIARIA 

Columna I P1&P3 CALIDAD DIARIA 

Columna J MTS^2 DEL DÍA 



Tercer tabla 

L:M

Columna L MES GARANTÍAS 

Columna M GARANTIAS 



Cuarta tabla 

O:P

Columna O MODELO DE PRUEBA

Columna P HORNO PARA PRUEBAS 



Quinta tabla 

R:S

Columna R MODELOS AUTORIZADOS

Columna S HORNO DE MODELO AUTORIZADOS 



Sexta tabla 

U:X

Columna U FECHA CUMPLIMIENTO A TONO

Columna V %CUMPLIMIENTO A TONO P1

Columna W %CUMPLIMIENTO A TONO P3

Columna X %CUMPLIMIENTO A TONO ACUMULADO



Séptima tabla 

Z:AA

Columna Z DEFECTOS P1/P3

Columna AA % DEFECTOS P1/P3



Octava tabla 

AC:AD

Columna AC DEFECTOS P1

Columna AD % DEFECTO P1



Novena tabla 

AF:AG

Columna AF DEFECTOS P3

Columna AG % DEFECROS P3



Décima tabla 

AI:AJ

Columna AI ÁREA RESPONSABLE 

Columna AJ % DE ÁREA 

Diseñar un dashboard efectivo requiere transformar datos en historias visuales que permitan tomar decisiones rápidamente. A partir de tus 10 tablas, la información se puede estructurar en 4 secciones clave:

1. Indicadores Globales y Desempeño Operativo
Sección superior con tarjetas KPI de un vistazo y gráficos de evolución temporal.

Tarjetas KPI (Resumen):

MTS² Producidos (Hoy / Mes): Suma total de MTS 
2
  del día.

% Cumplimiento a Tono Acumulado: Métrica global de calidad en tono.

% Defectos Totales (P1&P3): Porcentaje general de merma o fallas.

Garantías Totales: Total de garantías acumuladas.

Gráfico de Línea - Tendencia de Calidad (Diaria vs. Mensual):

Eje X: Días / Meses.

Eje Y: Porcentaje de Calidad.

Series: P1, P3 y P1&P3 (comparando el desempeño diario con las metas anuales/mensuales).

Gráfico de Barras - Cumplimiento a Tono:

Eje X: Fecha.

Eje Y: % Cumplimiento (P1, P3 y Acumulado).

2. Análisis de Defectos y Calidad por Planta/Línea
Identificación de problemas principales mediante el Principio de Pareto (80/20).

Visual	Datos Utilizados	Propósito
Gráfico de Pareto / Barras	Defectos P1&P3, P1 y P3 (% DEFECTOS)	Identificar los defectos más recurrentes por planta/línea.
Gráfico de Rosca / Donut	Áreas Responsables (% DE ÁREA)	Distribuir el impacto de los defectos según la responsabilidad operativa.
Gráfico de Barras Comparativo	Garantías por Mes (MES GARANTÍAS)	Monitorizar el impacto post-producción y reclamos.
3. Matriz de Hornos y Modelos
Control de capacidad autorizada vs. modelos en prueba.

+------------------------------------+------------------------------------+
|        MODELOS AUTORIZADOS         |          MODELOS EN PRUEBA         |
|  (R:S - Modelo vs. Horno Autoriz.) |  (O:P - Modelo vs. Horno Prueba)   |
+------------------------------------+------------------------------------+
| • Tabla/Tarjetas con estado ok     | • Alerta visual de pruebas activas |
| • Filtro interactivo por horno     | • Asignación de horno para ensayos |
+------------------------------------+------------------------------------+
4. Filtros Interactivos (Slicers)
Para dinamizar el panel de control en herramientas como Power BI, Tableau o Excel, incluye los siguientes segmentadores:

Rango de Fechas / Mes: Afecta a las tablas de calidad diaria, cumplimiento a tono y garantías.

Planta / Línea: Selector para filtrar entre P1, P3 o P1&P3.

Horno: Filtro para auditar modelos autorizados y en prueba por equipo.

Herramienta recomendada para implementación
Power BI / Tableau: Opción ideal para automatizar las relaciones entre tablas utilizando conectores a tu base de datos o Excel.

Excel / Google Sheets: Puedes crear este panel consolidando las tablas en una pestaña de datos y usando Tablas Dinámicas + Gráficos Dinámicos + Segmentadores de Datos en una pestaña limpia de Dashboard.

lo tengo en un archivo de excel y solo hay que leer la hoja DASHBOARD, quiero el codigo app.py para hacerlo en streamlit, con el concepto de la imagen que te anexo a su vez te anexo el archivo de excel con la hoja DASHBOARD donde esta toda la info

Python
import pandas as pd
import openpyxl

excel_path = 'REPORTE P1 Y P3 AGOSTO 2026.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
print("Sheet names:", wb.sheetnames)

sheet = wb['DASHBOARD']
# Read rows to inspect non-empty data blocks
df = pd.DataFrame(sheet.values)
print("DF Shape:", df.shape)

# Let's inspect columns and sample non-null areas
for col_idx in range(df.shape[1]):
    col_vals = df.iloc[:, col_idx].dropna().tolist()
    if col_vals:
        print(f"Col {col_idx} (Letter {openpyxl.utils.get_column_letter(col_idx+1)}): {col_vals[:5]}")

Python
import os
print("Files in current dir:", os.listdir('.'))

Python
import pandas as pd
import openpyxl

excel_path = 'REPORTE P1 Y P3 AGOSTO 2026.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['DASHBOARD']
df = pd.DataFrame(sheet.values)
print("Shape:", df.shape)

for col_idx in range(df.shape[1]):
    col_vals = df.iloc[:, col_idx].dropna().tolist()
    if col_vals:
        print(f"Col {col_idx} (Letter {openpyxl.utils.get_column_letter(col_idx+1)}): {col_vals[:5]}")

Python
import pandas as pd
import openpyxl

excel_path = 'REPORTE P1 Y P3 AGOSTO 2026.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['DASHBOARD']
df = pd.DataFrame(sheet.values)

# Print headers / content of the blocks based on user specified columns:
# Tabla 1: A:D (cols 0:4)
# Tabla 2: F:J (cols 5:10)
# Tabla 3: L:M (cols 11:13)
# Tabla 4: O:P (cols 14:16)
# Tabla 5: R:S (cols 17:19)
# Tabla 6: U:X (cols 20:24)
# Tabla 7: Z:AA (cols 25:27)
# Tabla 8: AC:AD (cols 28:30)
# Tabla 9: AF:AG (cols 31:33)
# Tabla 10: AI:AJ (cols 34:36)

col_ranges = {
    "T1_Mes_Calidad_Anual": (0, 4),
    "T2_Dia_Calidad_Diaria": (5, 10),
    "T3_Mes_Garantias": (11, 13),
    "T4_Modelo_Prueba": (14, 16),
    "T5_Modelos_Autorizados": (17, 19),
    "T6_Cumplimiento_Tono": (20, 24),
    "T7_Defectos_P1_P3": (25, 27),
    "T8_Defectos_P1": (28, 30),
    "T9_Defectos_P3": (31, 33),
    "T10_Area_Responsable": (34, 36)
}

for name, (start, end) in col_ranges.items():
    sub_df = df.iloc[:, start:end].dropna(how='all')
    print(f"=== {name} ===")
    print(sub_df.head(10))
    print("\n")

Python
import openpyxl
import pandas as pd

excel_path = 'REPORTE P1 Y P3 AGOSTO 2026.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['DASHBOARD']
df = pd.DataFrame(sheet.values)

print("Columns count:", df.shape[1])
print("Rows count:", df.shape[0])

for col in range(min(40, df.shape[1])):
    non_nulls = df.iloc[:, col].dropna()
    if len(non_nulls) > 0:
        col_letter = openpyxl.utils.get_column_letter(col+1)
        print(f"Col {col_letter} ({col}): first 3 non-null -> {non_nulls.head(3).tolist()}")

Para construir tu dashboard en Streamlit emulando la estética moderna y estructurada de tu imagen de referencia (modo oscuro profesional, indicadores tipo card con KPIs/metas, gráficos interactivos con Plotly, Pareto de defectos, tabla de modelos/áreas y menú de navegación lateral), aquí tienes la solución completa en Python.

Código completo: app.py
Python
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

# -----------------------------------------------------------------------------
# 1. CONFIGURACIÓN DE PÁGINA Y ESTILOS CSS (Estética basada en la imagen)
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="Dashboard - Sistema de Calidad",
    page_icon="🟩",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Inyección de CSS para modo oscuro, tarjetas KPI y diseño pulido
st.markdown("""
<style>
    /* Fondo principal y tipografía */
    .main {
        background-color: #0E131F;
        color: #FFFFFF;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    .block-container {
        padding-top: 1.5rem;
        padding-bottom: 2rem;
        padding-left: 2rem;
        padding-right: 2rem;
    }
    /* Sidebar */
    [data-testid="stSidebar"] {
        background-color: #080B11;
        border-right: 1px solid #1E293B;
    }
    
    /* Header superior */
    .dashboard-header {
        background-color: #131B2E;
        padding: 15px 25px;
        border-radius: 10px;
        border-left: 5px solid #10B981;
        margin-bottom: 20px;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
    .header-title {
        font-size: 24px;
        font-weight: 700;
        color: #FFFFFF;
        margin: 0;
    }
    .header-subtitle {
        font-size: 13px;
        color: #94A3B8;
        margin: 0;
        letter-spacing: 1px;
    }

    /* Cards KPI */
    .kpi-card {
        background-color: #131B2E;
        border: 1px solid #1E293B;
        border-radius: 12px;
        padding: 16px;
        text-align: center;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.3);
    }
    .kpi-title {
        font-size: 11px;
        font-weight: 700;
        color: #94A3B8;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-bottom: 8px;
    }
    .kpi-value {
        font-size: 26px;
        font-weight: 800;
        margin-bottom: 4px;
    }
    .kpi-subtext {
        font-size: 11px;
        color: #64748B;
    }

    /* Contenedor de Gráficos / Secciones */
    .section-box {
        background-color: #131B2E;
        border: 1px solid #1E293B;
        border-radius: 12px;
        padding: 18px;
        margin-bottom: 15px;
    }
    .section-title {
        font-size: 14px;
        font-weight: 700;
        color: #E2E8F0;
        margin-bottom: 15px;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
</style>
""", unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# 2. CARGA DE DATOS DESDE EXCEL ('DASHBOARD')
# -----------------------------------------------------------------------------
@st.cache_data
def load_data(file_path):
    # Carga de la hoja DASHBOARD
    xls = pd.ExcelFile(file_path)
    df_raw = pd.read_excel(xls, sheet_name='DASHBOARD')
    
    # Extracción y limpieza por rangos de columnas segun especificación
    # Tabla 1: A:D (Calidad Anual)
    t1 = df_raw.iloc[:, 0:4].dropna(how='all')
    t1.columns = ['MES', 'P1_ANUAL', 'P3_ANUAL', 'P1_P3_ANUAL']
    
    # Tabla 2: F:J (Calidad Diaria)
    t2 = df_raw.iloc[:, 5:10].dropna(how='all')
    t2.columns = ['DIA', 'P1_DIARIA', 'P3_DIARIA', 'P1_P3_DIARIA', 'MTS2_DIA']
    
    # Tabla 3: L:M (Garantías)
    t3 = df_raw.iloc[:, 11:13].dropna(how='all')
    t3.columns = ['MES_GARANTIAS', 'GARANTIAS']
    
    # Tabla 4: O:P (Modelos de Prueba)
    t4 = df_raw.iloc[:, 14:16].dropna(how='all')
    t4.columns = ['MODELO_PRUEBA', 'HORNO_PRUEBAS']
    
    # Tabla 5: R:S (Modelos Autorizados)
    t5 = df_raw.iloc[:, 17:19].dropna(how='all')
    t5.columns = ['MODELOS_AUTORIZADOS', 'HORNO_AUTORIZADOS']
    
    # Tabla 6: U:X (Cumplimiento a Tono)
    t6 = df_raw.iloc[:, 20:24].dropna(how='all')
    t6.columns = ['FECHA', 'CUMP_P1', 'CUMP_P3', 'CUMP_ACUMULADO']
    
    # Tabla 7: Z:AA (Defectos P1/P3)
    t7 = df_raw.iloc[:, 25:27].dropna(how='all')
    t7.columns = ['DEFECTO', 'PORC_DEFECTO']
    
    # Tabla 8: AC:AD (Defectos P1)
    t8 = df_raw.iloc[:, 28:30].dropna(how='all')
    t8.columns = ['DEFECTO_P1', 'PORC_DEFECTO_P1']
    
    # Tabla 9: AF:AG (Defectos P3)
    t9 = df_raw.iloc[:, 31:33].dropna(how='all')
    t9.columns = ['DEFECTO_P3', 'PORC_DEFECTO_P3']
    
    # Tabla 10: AI:AJ (Área Responsable)
    t10 = df_raw.iloc[:, 34:36].dropna(how='all')
    t10.columns = ['AREA_RESPONSABLE', 'PORC_AREA']
    
    return t1, t2, t3, t4, t5, t6, t7, t8, t9, t10

# Cargar archivo (Asegúrate de tener el archivo en el mismo directorio)
EXCEL_FILE = 'REPORTE P1 Y P3 AGOSTO 2026.xlsx'

try:
    t1, t2, t3, t4, t5, t6, t7, t8, t9, t10 = load_data(EXCEL_FILE)
except Exception as e:
    st.error(f"Error al cargar el archivo de Excel: {e}")
    st.stop()


# -----------------------------------------------------------------------------
# 3. BARRA LATERAL (NAVBAR ESTILO NAVEGACIÓN)
# -----------------------------------------------------------------------------
with st.sidebar:
    st.markdown("### 🟩 **SISTEMA DE CALIDAD**")
    st.caption("PISO CERÁMICO P1 & P3")
    st.markdown("---")
    
    menu = st.radio(
        "NAVEGACIÓN",
        ["RESUMEN", "INDICADORES", "DEFECTOS", "MODELOS Y HORNOS"],
        icons=["house", "bar-chart-line", "exclamation-triangle", "gear"]
    )
    
    st.markdown("---")
    st.markdown("### 🔍 **Filtros Globales**")
    
    # Filtro de Línea/Planta
    planta_sel = st.selectbox("Seleccionar Planta / Línea", ["Todas (P1 & P3)", "Planta 1 (P1)", "Planta 3 (P3)"])
    
    st.markdown("---")
    st.caption("Última Actualización: Agosto 2026")


# -----------------------------------------------------------------------------
# 4. ENCABEZADO PRINCIPAL
# -----------------------------------------------------------------------------
st.markdown("""
<div class="dashboard-header">
    <div>
        <div class="header-title">DASHBOARD - SISTEMA DE CALIDAD</div>
        <div class="header-subtitle">MONITORIZACIÓN Y CONTROL DE PRODUCCIÓN CERÁMICA</div>
    </div>
</div>
""", unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# 5. VISTA 1: RESUMEN GENERAL (INSPIRADO EN LA IMAGEN)
# -----------------------------------------------------------------------------
if menu == "RESUMEN":
    
    # --- FILA 1: CARDS DE KPI DE ALTO NIVEL ---
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    
    # Cálculo de métricas promedio / acumuladas
    val_p1_p3 = t2['P1_P3_DIARIA'].mean() if not t2.empty else 0
    val_p1 = t2['P1_DIARIA'].mean() if not t2.empty else 0
    val_p3 = t2['P3_DIARIA'].mean() if not t2.empty else 0
    val_mts = t2['MTS2_DIA'].sum() if not t2.empty else 0
    val_tono = t6['CUMP_ACUMULADO'].iloc[-1] if not t6.empty else 0
    val_garantias = t3['GARANTIAS'].sum() if not t3.empty else 0

    with c1:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">CALIDAD P1 & P3</div>
            <div class="kpi-value" style="color: #10B981;">{val_p1_p3*100:.1f}%</div>
            <div class="kpi-subtext">Meta ≥ 95.0%</div>
        </div>
        """, unsafe_allow_html=True)
        
    with c2:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">CALIDAD P1</div>
            <div class="kpi-value" style="color: #3B82F6;">{val_p1*100:.1f}%</div>
            <div class="kpi-subtext">Meta ≥ 95.0%</div>
        </div>
        """, unsafe_allow_html=True)

    with c3:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">CALIDAD P3</div>
            <div class="kpi-value" style="color: #F59E0B;">{val_p3*100:.1f}%</div>
            <div class="kpi-subtext">Meta ≥ 95.0%</div>
        </div>
        """, unsafe_allow_html=True)

    with c4:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">PRODUCCIÓN M²</div>
            <div class="kpi-value" style="color: #6366F1;">{val_mts:,.0f}</div>
            <div class="kpi-subtext">MTS² Acumulados</div>
        </div>
        """, unsafe_allow_html=True)

    with c5:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">CUMP. TONO</div>
            <div class="kpi-value" style="color: #EC4899;">{val_tono*100:.1f}%</div>
            <div class="kpi-subtext">Meta ≥ 98.0%</div>
        </div>
        """, unsafe_allow_html=True)

    with c6:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">GARANTÍAS</div>
            <div class="kpi-value" style="color: #EF4444;">{val_garantias:,.0f}</div>
            <div class="kpi-subtext">Reclamos Mes</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # --- FILA 2: PARETO DE DEFECTOS Y DISTRIBUCIÓN POR ÁREA ---
    col_g1, col_g2 = st.columns([1.2, 1])

    with col_g1:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">PARETO DE DEFECTOS PRINCIPALES (P1 & P3)</div>', unsafe_allow_html=True)
        
        # Selección de datos según filtro
        if planta_sel == "Planta 1 (P1)":
            df_def = t8.rename(columns={'DEFECTO_P1': 'DEFECTO', 'PORC_DEFECTO_P1': 'PORC_DEFECTO'})
        elif planta_sel == "Planta 3 (P3)":
            df_def = t9.rename(columns={'DEFECTO_P3': 'DEFECTO', 'PORC_DEFECTO_P3': 'PORC_DEFECTO'})
        else:
            df_def = t7

        if not df_def.empty:
            df_def = df_def.sort_values(by='PORC_DEFECTO', ascending=True)
            
            fig_def = px.bar(
                df_def,
                x='PORC_DEFECTO',
                y='DEFECTO',
                orientation='h',
                text=df_def['PORC_DEFECTO'].apply(lambda x: f"{x*100:.1f}%" if x <= 1 else f"{x:.1f}%"),
                color_discrete_sequence=['#3B82F6']
            )
            fig_def.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#94A3B8'),
                margin=dict(l=10, r=10, t=10, b=10),
                xaxis=dict(showgrid=False, visible=False),
                yaxis=dict(showgrid=False)
            )
            st.plotly_chart(fig_def, use_container_width=True)
        else:
            st.info("No hay datos de defectos disponibles.")
        st.markdown('</div>', unsafe_allow_html=True)

    with col_g2:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">DISTRIBUCIÓN DE DEFECTOS POR ÁREA RESPONSABLE</div>', unsafe_allow_html=True)
        
        if not t10.empty:
            fig_donut = px.pie(
                t10,
                names='AREA_RESPONSABLE',
                values='PORC_AREA',
                hole=0.6,
                color_discrete_sequence=px.colors.qualitative.Set3
            )
            fig_donut.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#94A3B8'),
                margin=dict(l=10, r=10, t=10, b=10),
                legend=dict(orientation="h", yanchor="bottom", y=-0.2)
            )
            st.plotly_chart(fig_donut, use_container_width=True)
        else:
            st.info("No hay datos de áreas responsables.")
        st.markdown('</div>', unsafe_allow_html=True)

    # --- FILA 3: TENDENCIA DE CALIDAD DIARIA ---
    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">EVOLUCIÓN DIARIA DE LA CALIDAD (%)</div>', unsafe_allow_html=True)
    
    if not t2.empty:
        fig_line = go.Figure()
        
        if planta_sel in ["Todas (P1 & P3)", "Planta 1 (P1)"]:
            fig_line.add_trace(go.Scatter(
                x=t2['DIA'], y=t2['P1_DIARIA']*100,
                mode='lines+markers', name='P1 Calidad',
                line=dict(color='#3B82F6', width=2)
            ))
            
        if planta_sel in ["Todas (P1 & P3)", "Planta 3 (P3)"]:
            fig_line.add_trace(go.Scatter(
                x=t2['DIA'], y=t2['P3_DIARIA']*100,
                mode='lines+markers', name='P3 Calidad',
                line=dict(color='#F59E0B', width=2)
            ))

        fig_line.add_trace(go.Scatter(
            x=t2['DIA'], y=t2['P1_P3_DIARIA']*100,
            mode='lines+markers', name='P1&P3 Global',
            line=dict(color='#10B981', width=3, dash='dash')
        ))

        fig_line.update_layout(
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(color='#94A3B8'),
            margin=dict(l=10, r=10, t=10, b=10),
            xaxis=dict(showgrid=True, gridcolor='#1E293B', title="Día"),
            yaxis=dict(showgrid=True, gridcolor='#1E293B', title="% Calidad"),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
        )
        st.plotly_chart(fig_line, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# 6. VISTA 2: INDICADORES TEMPORALES (ANUAL / GARANTÍAS)
# -----------------------------------------------------------------------------
elif menu == "INDICADORES":
    col_ind1, col_ind2 = st.columns(2)
    
    with col_ind1:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">HISTÓRICO DE CALIDAD MES A MES</div>', unsafe_allow_html=True)
        if not t1.empty:
            fig_anual = px.line(
                t1, x='MES', y=['P1_ANUAL', 'P3_ANUAL', 'P1_P3_ANUAL'],
                markers=True,
                color_discrete_map={'P1_ANUAL': '#3B82F6', 'P3_ANUAL': '#F59E0B', 'P1_P3_ANUAL': '#10B981'}
            )
            fig_anual.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#94A3B8'),
                xaxis=dict(showgrid=True, gridcolor='#1E293B'),
                yaxis=dict(showgrid=True, gridcolor='#1E293B')
            )
            st.plotly_chart(fig_anual, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
        
    with col_ind2:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">GARANTÍAS RECLAMADAS POR MES</div>', unsafe_allow_html=True)
        if not t3.empty:
            fig_gar = px.bar(
                t3, x='MES_GARANTIAS', y='GARANTIAS',
                text='GARANTIAS',
                color_discrete_sequence=['#EF4444']
            )
            fig_gar.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#94A3B8'),
                xaxis=dict(showgrid=False),
                yaxis=dict(showgrid=True, gridcolor='#1E293B')
            )
            st.plotly_chart(fig_gar, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # Cumplimiento a Tono
    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">CUMPLIMIENTO DE TONO EN PRODUCCIÓN</div>', unsafe_allow_html=True)
    if not t6.empty:
        fig_tono = px.bar(
            t6, x='FECHA', y=['CUMP_P1', 'CUMP_P3', 'CUMP_ACUMULADO'],
            barmode='group',
            color_discrete_sequence=['#3B82F6', '#F59E0B', '#10B981']
        )
        fig_tono.update_layout(
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(color='#94A3B8'),
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor='#1E293B')
        )
        st.plotly_chart(fig_tono, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# 7. VISTA 3: ANÁLISIS DETALLADO DE DEFECTOS
# -----------------------------------------------------------------------------
elif menu == "DEFECTOS":
    col_d1, col_d2 = st.columns(2)
    
    with col_d1:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">DEFECTOS EN PLANTA 1 (P1)</div>', unsafe_allow_html=True)
        st.dataframe(t8, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with col_d2:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">DEFECTOS EN PLANTA 3 (P3)</div>', unsafe_allow_html=True)
        st.dataframe(t9, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">DEFECTOS GLOBALES Y RESPONSABILIDAD DE ÁREA</div>', unsafe_allow_html=True)
    c_a, c_b = st.columns(2)
    with c_a:
        st.subheader("Defectos Consolidados (P1 & P3)")
        st.dataframe(t7, use_container_width=True)
    with c_b:
        st.subheader("% Defectos por Área Responsable")
        st.dataframe(t10, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# 8. VISTA 4: CONTROL DE MODELOS Y HORNOS
# -----------------------------------------------------------------------------
elif menu == "MODELOS Y HORNOS":
    col_m1, col_m2 = st.columns(2)

    with col_m1:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">🧪 MODELOS DE PRUEBA EN HORNO</div>', unsafe_allow_html=True)
        st.dataframe(t4, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with col_m2:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">✅ MODELOS AUTORIZADOS</div>', unsafe_allow_html=True)
        st.dataframe(t5, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
Pasos para ejecutar la aplicación:
Instalar las dependencias necesarias (abre tu terminal o consola):

Bash
pip install streamlit pandas openpyxl plotly
Organización de archivos:
Asegúrate de tener el archivo de Python y tu reporte de Excel en la misma carpeta:

Plaintext
📁 MiProyecto/
├── 📄 app.py
└── 📊 REPORTE P1 Y P3 AGOSTO 2026.xlsx
Ejecutar la aplicación de Streamlit:

Bash
streamlit run app.py
Aspectos clave del diseño:
UI/UX alineada a la imagen: Tema oscuro (#0E131F) con tarjetas KPI enriquecidas en la parte superior, bordes redondeados y tipografía pulida.

Procesamiento de 10 Tablas: Lee automáticamente desde la pestaña DASHBOARD dividida por las columnas A:D, F:J, L:M, O:P, R:S, U:X, Z:AA, AC:AD, AF:AG y AI:AJ.

Interactividad con Plotly: Gráficos adaptativos (Pareto horizontal, Donut chart, series de tiempo y barras comparativas) que responden dinámicamente a los filtros de planta.

todo junto o el descargable

Aquí tienes listo el archivo de código descargable listo para ejecutar:

Your Python file is ready

Icono de PY
app
 PY 
Abierto
Instrucciones rápidas de instalación y ejecución:
Guardar archivos: Coloca el archivo app.py descargado en la misma carpeta donde se encuentra tu Excel REPORTE P1 Y P3 AGOSTO 2026.xlsx.

Instalar dependencias (si aún no las tienes):

Bash
pip install streamlit pandas openpyxl plotly
Ejecutar el Dashboard:

Bash
streamlit run app.py



Gemini es una IA y puede cometer errores.

Analizando
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

# -----------------------------------------------------------------------------
# 1. CONFIGURACIÓN DE PÁGINA Y ESTILOS CSS (Estética basada en la imagen)
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="Dashboard - Sistema de Calidad",
    page_icon="🟩",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Inyección de CSS para modo oscuro, tarjetas KPI y diseño pulido
st.markdown("""
<style>
    /* Fondo principal y tipografía */
    .main {
        background-color: #0E131F;
        color: #FFFFFF;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    .block-container {
        padding-top: 1.5rem;
        padding-bottom: 2rem;
        padding-left: 2rem;
        padding-right: 2rem;
    }
    /* Sidebar */
    [data-testid="stSidebar"] {
        background-color: #080B11;
        border-right: 1px solid #1E293B;
    }
    
    /* Header superior */
    .dashboard-header {
        background-color: #131B2E;
        padding: 15px 25px;
        border-radius: 10px;
        border-left: 5px solid #10B981;
        margin-bottom: 20px;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
    .header-title {
        font-size: 24px;
        font-weight: 700;
        color: #FFFFFF;
        margin: 0;
    }
    .header-subtitle {
        font-size: 13px;
        color: #94A3B8;
        margin: 0;
        letter-spacing: 1px;
    }

    /* Cards KPI */
    .kpi-card {
        background-color: #131B2E;
        border: 1px solid #1E293B;
        border-radius: 12px;
        padding: 16px;
        text-align: center;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.3);
    }
    .kpi-title {
        font-size: 11px;
        font-weight: 700;
        color: #94A3B8;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-bottom: 8px;
    }
    .kpi-value {
        font-size: 26px;
        font-weight: 800;
        margin-bottom: 4px;
    }
    .kpi-subtext {
        font-size: 11px;
        color: #64748B;
    }

    /* Contenedor de Gráficos / Secciones */
    .section-box {
        background-color: #131B2E;
        border: 1px solid #1E293B;
        border-radius: 12px;
        padding: 18px;
        margin-bottom: 15px;
    }
    .section-title {
        font-size: 14px;
        font-weight: 700;
        color: #E2E8F0;
        margin-bottom: 15px;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
</style>
""", unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# 2. CARGA DE DATOS DESDE EXCEL ('DASHBOARD')
# -----------------------------------------------------------------------------
@st.cache_data
def load_data(file_path):
    # Carga de la hoja DASHBOARD
    xls = pd.ExcelFile(file_path)
    df_raw = pd.read_excel(xls, sheet_name='DASHBOARD')
    
    # Extracción y limpieza por rangos de columnas segun especificación
    # Tabla 1: A:D (Calidad Anual)
    t1 = df_raw.iloc[:, 0:4].dropna(how='all')
    t1.columns = ['MES', 'P1_ANUAL', 'P3_ANUAL', 'P1_P3_ANUAL']
    
    # Tabla 2: F:J (Calidad Diaria)
    t2 = df_raw.iloc[:, 5:10].dropna(how='all')
    t2.columns = ['DIA', 'P1_DIARIA', 'P3_DIARIA', 'P1_P3_DIARIA', 'MTS2_DIA']
    
    # Tabla 3: L:M (Garantías)
    t3 = df_raw.iloc[:, 11:13].dropna(how='all')
    t3.columns = ['MES_GARANTIAS', 'GARANTIAS']
    
    # Tabla 4: O:P (Modelos de Prueba)
    t4 = df_raw.iloc[:, 14:16].dropna(how='all')
    t4.columns = ['MODELO_PRUEBA', 'HORNO_PRUEBAS']
    
    # Tabla 5: R:S (Modelos Autorizados)
    t5 = df_raw.iloc[:, 17:19].dropna(how='all')
    t5.columns = ['MODELOS_AUTORIZADOS', 'HORNO_AUTORIZADOS']
    
    # Tabla 6: U:X (Cumplimiento a Tono)
    t6 = df_raw.iloc[:, 20:24].dropna(how='all')
    t6.columns = ['FECHA', 'CUMP_P1', 'CUMP_P3', 'CUMP_ACUMULADO']
    
    # Tabla 7: Z:AA (Defectos P1/P3)
    t7 = df_raw.iloc[:, 25:27].dropna(how='all')
    t7.columns = ['DEFECTO', 'PORC_DEFECTO']
    
    # Tabla 8: AC:AD (Defectos P1)
    t8 = df_raw.iloc[:, 28:30].dropna(how='all')
    t8.columns = ['DEFECTO_P1', 'PORC_DEFECTO_P1']
    
    # Tabla 9: AF:AG (Defectos P3)
    t9 = df_raw.iloc[:, 31:33].dropna(how='all')
    t9.columns = ['DEFECTO_P3', 'PORC_DEFECTO_P3']
    
    # Tabla 10: AI:AJ (Área Responsable)
    t10 = df_raw.iloc[:, 34:36].dropna(how='all')
    t10.columns = ['AREA_RESPONSABLE', 'PORC_AREA']
    
    return t1, t2, t3, t4, t5, t6, t7, t8, t9, t10

# Cargar archivo (Asegúrate de tener el archivo en el mismo directorio)
EXCEL_FILE = 'REPORTE P1 Y P3 AGOSTO 2026.xlsx'

try:
    t1, t2, t3, t4, t5, t6, t7, t8, t9, t10 = load_data(EXCEL_FILE)
except Exception as e:
    st.error(f"Error al cargar el archivo de Excel: {e}")
    st.stop()


# -----------------------------------------------------------------------------
# 3. BARRA LATERAL (NAVBAR ESTILO NAVEGACIÓN)
# -----------------------------------------------------------------------------
with st.sidebar:
    st.markdown("### 🟩 **SISTEMA DE CALIDAD**")
    st.caption("PISO CERÁMICO P1 & P3")
    st.markdown("---")
    
    menu = st.radio(
        "NAVEGACIÓN",
        ["RESUMEN", "INDICADORES", "DEFECTOS", "MODELOS Y HORNOS"],
        icons=["house", "bar-chart-line", "exclamation-triangle", "gear"]
    )
    
    st.markdown("---")
    st.markdown("### 🔍 **Filtros Globales**")
    
    # Filtro de Línea/Planta
    planta_sel = st.selectbox("Seleccionar Planta / Línea", ["Todas (P1 & P3)", "Planta 1 (P1)", "Planta 3 (P3)"])
    
    st.markdown("---")
    st.caption("Última Actualización: Agosto 2026")


# -----------------------------------------------------------------------------
# 4. ENCABEZADO PRINCIPAL
# -----------------------------------------------------------------------------
st.markdown("""
<div class="dashboard-header">
    <div>
        <div class="header-title">DASHBOARD - SISTEMA DE CALIDAD</div>
        <div class="header-subtitle">MONITORIZACIÓN Y CONTROL DE PRODUCCIÓN CERÁMICA</div>
    </div>
</div>
""", unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# 5. VISTA 1: RESUMEN GENERAL (INSPIRADO EN LA IMAGEN)
# -----------------------------------------------------------------------------
if menu == "RESUMEN":
    
    # --- FILA 1: CARDS DE KPI DE ALTO NIVEL ---
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    
    # Cálculo de métricas promedio / acumuladas
    val_p1_p3 = t2['P1_P3_DIARIA'].mean() if not t2.empty else 0
    val_p1 = t2['P1_DIARIA'].mean() if not t2.empty else 0
    val_p3 = t2['P3_DIARIA'].mean() if not t2.empty else 0
    val_mts = t2['MTS2_DIA'].sum() if not t2.empty else 0
    val_tono = t6['CUMP_ACUMULADO'].iloc[-1] if not t6.empty else 0
    val_garantias = t3['GARANTIAS'].sum() if not t3.empty else 0

    with c1:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">CALIDAD P1 & P3</div>
            <div class="kpi-value" style="color: #10B981;">{val_p1_p3*100:.1f}%</div>
            <div class="kpi-subtext">Meta ≥ 95.0%</div>
        </div>
        """, unsafe_allow_html=True)
        
    with c2:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">CALIDAD P1</div>
            <div class="kpi-value" style="color: #3B82F6;">{val_p1*100:.1f}%</div>
            <div class="kpi-subtext">Meta ≥ 95.0%</div>
        </div>
        """, unsafe_allow_html=True)

    with c3:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">CALIDAD P3</div>
            <div class="kpi-value" style="color: #F59E0B;">{val_p3*100:.1f}%</div>
            <div class="kpi-subtext">Meta ≥ 95.0%</div>
        </div>
        """, unsafe_allow_html=True)

    with c4:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">PRODUCCIÓN M²</div>
            <div class="kpi-value" style="color: #6366F1;">{val_mts:,.0f}</div>
            <div class="kpi-subtext">MTS² Acumulados</div>
        </div>
        """, unsafe_allow_html=True)

    with c5:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">CUMP. TONO</div>
            <div class="kpi-value" style="color: #EC4899;">{val_tono*100:.1f}%</div>
            <div class="kpi-subtext">Meta ≥ 98.0%</div>
        </div>
        """, unsafe_allow_html=True)

    with c6:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">GARANTÍAS</div>
            <div class="kpi-value" style="color: #EF4444;">{val_garantias:,.0f}</div>
            <div class="kpi-subtext">Reclamos Mes</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # --- FILA 2: PARETO DE DEFECTOS Y DISTRIBUCIÓN POR ÁREA ---
    col_g1, col_g2 = st.columns([1.2, 1])

    with col_g1:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">PARETO DE DEFECTOS PRINCIPALES (P1 & P3)</div>', unsafe_allow_html=True)
        
        # Selección de datos según filtro
        if planta_sel == "Planta 1 (P1)":
            df_def = t8.rename(columns={'DEFECTO_P1': 'DEFECTO', 'PORC_DEFECTO_P1': 'PORC_DEFECTO'})
        elif planta_sel == "Planta 3 (P3)":
            df_def = t9.rename(columns={'DEFECTO_P3': 'DEFECTO', 'PORC_DEFECTO_P3': 'PORC_DEFECTO'})
        else:
            df_def = t7

        if not df_def.empty:
            df_def = df_def.sort_values(by='PORC_DEFECTO', ascending=True)
            
            fig_def = px.bar(
                df_def,
                x='PORC_DEFECTO',
                y='DEFECTO',
                orientation='h',
                text=df_def['PORC_DEFECTO'].apply(lambda x: f"{x*100:.1f}%" if x <= 1 else f"{x:.1f}%"),
                color_discrete_sequence=['#3B82F6']
            )
            fig_def.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#94A3B8'),
                margin=dict(l=10, r=10, t=10, b=10),
                xaxis=dict(showgrid=False, visible=False),
                yaxis=dict(showgrid=False)
            )
            st.plotly_chart(fig_def, use_container_width=True)
        else:
            st.info("No hay datos de defectos disponibles.")
        st.markdown('</div>', unsafe_allow_html=True)

    with col_g2:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">DISTRIBUCIÓN DE DEFECTOS POR ÁREA RESPONSABLE</div>', unsafe_allow_html=True)
        
        if not t10.empty:
            fig_donut = px.pie(
                t10,
                names='AREA_RESPONSABLE',
                values='PORC_AREA',
                hole=0.6,
                color_discrete_sequence=px.colors.qualitative.Set3
            )
            fig_donut.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#94A3B8'),
                margin=dict(l=10, r=10, t=10, b=10),
                legend=dict(orientation="h", yanchor="bottom", y=-0.2)
            )
            st.plotly_chart(fig_donut, use_container_width=True)
        else:
            st.info("No hay datos de áreas responsables.")
        st.markdown('</div>', unsafe_allow_html=True)

    # --- FILA 3: TENDENCIA DE CALIDAD DIARIA ---
    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">EVOLUCIÓN DIARIA DE LA CALIDAD (%)</div>', unsafe_allow_html=True)
    
    if not t2.empty:
        fig_line = go.Figure()
        
        if planta_sel in ["Todas (P1 & P3)", "Planta 1 (P1)"]:
            fig_line.add_trace(go.Scatter(
                x=t2['DIA'], y=t2['P1_DIARIA']*100,
                mode='lines+markers', name='P1 Calidad',
                line=dict(color='#3B82F6', width=2)
            ))
            
        if planta_sel in ["Todas (P1 & P3)", "Planta 3 (P3)"]:
            fig_line.add_trace(go.Scatter(
                x=t2['DIA'], y=t2['P3_DIARIA']*100,
                mode='lines+markers', name='P3 Calidad',
                line=dict(color='#F59E0B', width=2)
            ))

        fig_line.add_trace(go.Scatter(
            x=t2['DIA'], y=t2['P1_P3_DIARIA']*100,
            mode='lines+markers', name='P1&P3 Global',
            line=dict(color='#10B981', width=3, dash='dash')
        ))

        fig_line.update_layout(
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(color='#94A3B8'),
            margin=dict(l=10, r=10, t=10, b=10),
            xaxis=dict(showgrid=True, gridcolor='#1E293B', title="Día"),
            yaxis=dict(showgrid=True, gridcolor='#1E293B', title="% Calidad"),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
        )
        st.plotly_chart(fig_line, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# 6. VISTA 2: INDICADORES TEMPORALES (ANUAL / GARANTÍAS)
# -----------------------------------------------------------------------------
elif menu == "INDICADORES":
    col_ind1, col_ind2 = st.columns(2)
    
    with col_ind1:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">HISTÓRICO DE CALIDAD MES A MES</div>', unsafe_allow_html=True)
        if not t1.empty:
            fig_anual = px.line(
                t1, x='MES', y=['P1_ANUAL', 'P3_ANUAL', 'P1_P3_ANUAL'],
                markers=True,
                color_discrete_map={'P1_ANUAL': '#3B82F6', 'P3_ANUAL': '#F59E0B', 'P1_P3_ANUAL': '#10B981'}
            )
            fig_anual.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#94A3B8'),
                xaxis=dict(showgrid=True, gridcolor='#1E293B'),
                yaxis=dict(showgrid=True, gridcolor='#1E293B')
            )
            st.plotly_chart(fig_anual, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
        
    with col_ind2:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">GARANTÍAS RECLAMADAS POR MES</div>', unsafe_allow_html=True)
        if not t3.empty:
            fig_gar = px.bar(
                t3, x='MES_GARANTIAS', y='GARANTIAS',
                text='GARANTIAS',
                color_discrete_sequence=['#EF4444']
            )
            fig_gar.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#94A3B8'),
                xaxis=dict(showgrid=False),
                yaxis=dict(showgrid=True, gridcolor='#1E293B')
            )
            st.plotly_chart(fig_gar, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # Cumplimiento a Tono
    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">CUMPLIMIENTO DE TONO EN PRODUCCIÓN</div>', unsafe_allow_html=True)
    if not t6.empty:
        fig_tono = px.bar(
            t6, x='FECHA', y=['CUMP_P1', 'CUMP_P3', 'CUMP_ACUMULADO'],
            barmode='group',
            color_discrete_sequence=['#3B82F6', '#F59E0B', '#10B981']
        )
        fig_tono.update_layout(
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(color='#94A3B8'),
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor='#1E293B')
        )
        st.plotly_chart(fig_tono, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# 7. VISTA 3: ANÁLISIS DETALLADO DE DEFECTOS
# -----------------------------------------------------------------------------
elif menu == "DEFECTOS":
    col_d1, col_d2 = st.columns(2)
    
    with col_d1:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">DEFECTOS EN PLANTA 1 (P1)</div>', unsafe_allow_html=True)
        st.dataframe(t8, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with col_d2:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">DEFECTOS EN PLANTA 3 (P3)</div>', unsafe_allow_html=True)
        st.dataframe(t9, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">DEFECTOS GLOBALES Y RESPONSABILIDAD DE ÁREA</div>', unsafe_allow_html=True)
    c_a, c_b = st.columns(2)
    with c_a:
        st.subheader("Defectos Consolidados (P1 & P3)")
        st.dataframe(t7, use_container_width=True)
    with c_b:
        st.subheader("% Defectos por Área Responsable")
        st.dataframe(t10, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# 8. VISTA 4: CONTROL DE MODELOS Y HORNOS
# -----------------------------------------------------------------------------
elif menu == "MODELOS Y HORNOS":
    col_m1, col_m2 = st.columns(2)

    with col_m1:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">🧪 MODELOS DE PRUEBA EN HORNO</div>', unsafe_allow_html=True)
        st.dataframe(t4, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with col_m2:
        st.markdown('<div class="section-box">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">✅ MODELOS AUTORIZADOS</div>', unsafe_allow_html=True)
        st.dataframe(t5, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
